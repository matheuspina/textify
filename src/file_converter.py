import os
import json
import csv
import yaml
import re
import unicodedata
import subprocess
from io import StringIO
from typing import Optional
from pathlib import Path

# Importações para diferentes formatos
try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import PyPDF2
    import pdfplumber
except ImportError:
    PyPDF2 = None
    pdfplumber = None

try:
    from pptx import Presentation
except ImportError:
    Presentation = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    from odf.opendocument import load
    from odf.text import P
    from odf.teletype import extractText
except ImportError:
    load = None
    P = None
    extractText = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from docx2txt import process as docx2txt_process
except ImportError:
    docx2txt_process = None

try:
    import xlrd
except ImportError:
    xlrd = None

class FileConverter:
    """Classe responsável por converter diferentes formatos de arquivo para texto"""
    
    def __init__(self):
        self._antiword_available = self._check_antiword_availability()
        
        self.supported_extensions = {
            '.docx': self._convert_docx,
            '.doc': self._convert_doc,
            '.xml': self._convert_xml,
            '.yml': self._convert_yaml,
            '.yaml': self._convert_yaml,
            '.xlsx': self._convert_xlsx,
            '.xls': self._convert_xls,
            '.csv': self._convert_csv,
            '.pdf': self._convert_pdf,
            '.txt': self._convert_txt,
            '.pptx': self._convert_pptx,
            '.ppt': self._convert_ppt,
            '.html': self._convert_html,
            '.htm': self._convert_html,
            '.odt': self._convert_odt,
            '.odp': self._convert_odp,
            '.ods': self._convert_ods,
            '.json': self._convert_json
        }
    
    def clean_text(self, text: str) -> str:
        """Limpa o texto removendo caracteres estranhos e formatação desnecessária"""
        if not text:
            return ""
        
        # Remove caracteres binários e de controle
        text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', text)
        
        # Remove sequências específicas do Word/DOC
        text = re.sub(r'bjbj[a-zA-Z0-9]+', '', text)
        text = re.sub(r'YgYg[a-zA-Z0-9]*', '', text)
        text = re.sub(r'~\$~\$~\$~\$~\$~\$\$', '', text)
        text = re.sub(r'CJOJQJ[^\s]*', '', text)
        text = re.sub(r'\^[a-zA-Z0-9`]+', '', text)
        text = re.sub(r'[a-zA-Z]\s[a-zA-Z]\s[a-zA-Z]\s[a-zA-Z]\s[a-zA-Z]', '', text)
        
        # Remove sequências de formatação específicas encontradas
        text = re.sub(r'Microsoft Office Word', '', text)
        text = re.sub(r'Word\.Document\.[0-9]+', '', text)
        text = re.sub(r'MSWordDoc', '', text)
        text = re.sub(r'Documento do Microsoft Word [0-9-]+', '', text)
        text = re.sub(r'Times New Roman|Arial|Calibri|Tahoma|Courier New|Wingdings|Cambria Math', '', text)
        
        # Remove códigos de formatação
        text = re.sub(r'\[[x\s]*[A-Za-z\s]*\]', '', text)
        text = re.sub(r'[A-Z]\s[a-z]\s[a-z]\s[a-z]\s[a-z]', '', text)
        
        # Remove sequências repetitivas de caracteres especiais
        text = re.sub(r'[^\w\sÀ-ÿ]{3,}', ' ', text)
        
        # Remove referências de arquivo e caminhos
        text = re.sub(r'[A-Z]:\\[^\s]*', '', text)
        text = re.sub(r'/[^\s]*\.(doc|docx|pdf|txt)', '', text)
        
        # Remove tags XML
        text = re.sub(r'<[^>]*>', '', text)
        
        # Remove códigos Unicode problemáticos
        text = re.sub(r'\\u[0-9a-fA-F]{4}', '', text)
        text = re.sub(r'\\x[0-9a-fA-F]{2}', '', text)
        
        # Remove URLs
        text = re.sub(r'https?://[^\s]+', '', text)
        
        # Normaliza espaços
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        
        # Processa linha por linha
        lines = text.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Remove caracteres especiais no início e fim
            line = re.sub(r'^[^\w\sÀ-ÿ]+', '', line)
            line = re.sub(r'[^\w\sÀ-ÿ]+$', '', line)
            line = line.strip()
            
            if not line:
                continue
                
            # Conta palavras legíveis (incluindo acentos)
            words = re.findall(r'\b[a-zA-ZÀ-ÿ]{2,}\b', line)
            
            # Mantém linhas com pelo menos 2 palavras legíveis OU pelo menos 10 caracteres úteis
            if len(words) >= 2 or (len(words) >= 1 and len(line) >= 10):
                cleaned_lines.append(line)
        
        result = '\n'.join(cleaned_lines)
        
        # Remove linhas muito curtas no final
        result = re.sub(r'\n[^\n]{1,3}\n', '\n', result)
        result = re.sub(r'\s+', ' ', result)
        
        return result.strip()
    
    async def convert_file(self, file_path: str, filename: str) -> str:
        """Converte um arquivo para texto baseado na extensão"""
        file_extension = Path(filename).suffix.lower()
        
        if file_extension not in self.supported_extensions:
            raise ValueError(f"Formato de arquivo não suportado: {file_extension}")
        
        converter_func = self.supported_extensions[file_extension]
        
        try:
            return await converter_func(file_path)
        except Exception as e:
            raise Exception(f"Erro ao converter arquivo {filename}: {str(e)}")
    
    async def _convert_docx(self, file_path: str) -> str:
        """Converte arquivo DOCX para texto"""
        if Document is None:
            raise ImportError("python-docx não está instalado")
        
        doc = Document(file_path)
        text_content = []
        
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_content.append(paragraph.text)
        
        return '\n'.join(text_content)
    
    async def _convert_xml(self, file_path: str) -> str:
        """Converte arquivo XML para texto"""
        if BeautifulSoup is None:
            raise ImportError("beautifulsoup4 não está instalado")
        
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        soup = BeautifulSoup(content, 'xml')
        return soup.get_text(separator='\n', strip=True)
    
    async def _convert_yaml(self, file_path: str) -> str:
        """Converte arquivo YAML para texto"""
        with open(file_path, 'r', encoding='utf-8') as file:
            data = yaml.safe_load(file)
        
        return json.dumps(data, indent=2, ensure_ascii=False)
    
    async def _convert_xlsx(self, file_path: str) -> str:
        """Converte arquivo XLSX para texto usando openpyxl."""
        if load_workbook is None:
            raise ImportError("openpyxl não está instalado. Não é possível converter arquivos .xlsx")

        try:
            workbook = load_workbook(file_path)
            text_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"=== Planilha: {sheet_name} ===")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        text_content.append(row_text)
            
            return '\n'.join(text_content)
        except Exception as e:
            raise Exception(f"Falha ao converter .xlsx com openpyxl: {e}")
    
    async def _convert_csv(self, file_path: str) -> str:
        """Converte arquivo CSV para texto"""
        text_content = []
        
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                text_content.append('\t'.join(row))
        
        return '\n'.join(text_content)
    
    async def _convert_pdf(self, file_path: str) -> str:
        """Converte arquivo PDF para texto"""
        text_content = []
        
        # Tenta usar pdfplumber primeiro (melhor para extração de texto)
        if pdfplumber is not None:
            try:
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            text_content.append(text)
                return '\n'.join(text_content)
            except Exception:
                pass
        
        # Fallback para PyPDF2
        if PyPDF2 is not None:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text = page.extract_text()
                    if text:
                        text_content.append(text)
            return '\n'.join(text_content)
        
        raise ImportError("Nenhuma biblioteca PDF está disponível")
    
    async def _convert_txt(self, file_path: str) -> str:
        """Converte arquivo TXT para texto"""
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    
    async def _extract_text_from_zip_xml(self, file_path: str, content_path_prefix: str, content_file: Optional[str] = None) -> str:
        """Extrai texto de arquivos baseados em zip/xml como .pptx e .odp"""
        try:
            import zipfile
            from xml.etree.ElementTree import fromstring
        except ImportError:
            raise ImportError("Módulos necessários (zipfile, xml) não estão disponíveis.")

        text_content = []
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                if content_file:
                    files_to_process = [content_file]
                else:
                    files_to_process = sorted([f for f in zf.namelist() if f.startswith(content_path_prefix)])

                for file_to_process in files_to_process:
                    if file_to_process in zf.namelist():
                        with zf.open(file_to_process) as xml_file:
                            xml_content = xml_file.read()
                            # Lógica para ODP
                            if 'content.xml' in file_to_process:
                                if load and P and extractText:
                                    try:
                                        # odfpy espera um arquivo ou nome de arquivo, então usamos BytesIO
                                        from io import BytesIO
                                        doc = load(BytesIO(xml_content))
                                        for p in doc.getElementsByType(P):
                                            text_content.append(extractText(p))
                                    except Exception as e:
                                        # Se odfpy falhar, tenta extração de texto genérica como fallback
                                        tree = fromstring(xml_content)
                                        for elem in tree.iter():
                                            if elem.text:
                                                text_content.append(elem.text.strip())
                                else:
                                    # Fallback se odfpy não estiver instalado
                                    tree = fromstring(xml_content)
                                    for elem in tree.iter():
                                        if elem.text:
                                            text_content.append(elem.text.strip())
                            # Lógica para PPTX
                            else:
                                tree = fromstring(xml_content)
                                for elem in tree.iter():
                                    if elem.text:
                                        text_content.append(elem.text.strip())
        except zipfile.BadZipFile:
            raise Exception("Arquivo não é um formato zip válido (e.g., .pptx)")
        except Exception as e:
            raise Exception(f"Falha ao extrair texto da apresentação: {e}")

        return '\n'.join(filter(None, text_content))

    async def _convert_ppt(self, file_path: str) -> str:
        """Converte arquivo PPT para texto usando LibreOffice para converter para PPTX primeiro"""
        import tempfile
        import shutil
        
        # Verifica se LibreOffice está disponível
        if not self._check_libreoffice_availability():
            raise Exception("LibreOffice não está disponível para conversão de arquivos .ppt")
        
        # Cria diretório temporário para a conversão
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                # Configurações adicionais para resolver problemas de permissão
                env = os.environ.copy()
                env['HOME'] = '/home/appuser'
                env['XDG_CONFIG_HOME'] = '/home/appuser/.config'
                env['XDG_CACHE_HOME'] = '/home/appuser/.cache'
                env['XDG_DATA_HOME'] = '/home/appuser/.local/share'
                
                # Converte PPT para PPTX usando LibreOffice com configurações específicas
                result = subprocess.run([
                    'libreoffice', 
                    '--headless', 
                    '--invisible',
                    '--nodefault',
                    '--nolockcheck',
                    '--nologo',
                    '--norestore',
                    '--convert-to', 'pptx',
                    '--outdir', temp_dir, 
                    file_path
                ], capture_output=True, text=True, timeout=120, env=env)
                
                if result.returncode != 0:
                    error_msg = result.stderr or result.stdout or "Erro desconhecido"
                    raise Exception(f"LibreOffice falhou na conversão: {error_msg}")
                
                # Encontra o arquivo PPTX gerado
                pptx_files = [f for f in os.listdir(temp_dir) if f.endswith('.pptx')]
                if not pptx_files:
                    raise Exception("Nenhum arquivo PPTX foi gerado pela conversão")
                
                pptx_path = os.path.join(temp_dir, pptx_files[0])
                
                # Converte o PPTX para texto usando o método existente
                return await self._convert_pptx(pptx_path)
                
            except subprocess.TimeoutExpired:
                raise Exception("Timeout na conversão do arquivo .ppt com LibreOffice")
            except Exception as e:
                raise Exception(f"Erro na conversão de .ppt: {str(e)}")

    async def _convert_pptx(self, file_path: str) -> str:
        """Converte arquivo PPTX para texto usando a extração de zip/xml."""
        return await self._extract_text_from_zip_xml(file_path, 'ppt/slides/slide')
    
    async def _convert_html(self, file_path: str) -> str:
        """Converte arquivo HTML para texto"""
        if BeautifulSoup is None:
            raise ImportError("beautifulsoup4 não está instalado")
        
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        soup = BeautifulSoup(content, 'html.parser')
        return soup.get_text(separator='\n', strip=True)
    
    async def _convert_odt(self, file_path: str) -> str:
        """Converte arquivo ODT para texto"""
        if load is None or extractText is None:
            raise ImportError("odfpy não está instalado")
        
        doc = load(file_path)
        text_content = []
        
        for paragraph in doc.getElementsByType(P):
            text = extractText(paragraph)
            if text.strip():
                text_content.append(text)
        
        return '\n'.join(text_content)
    
    async def _convert_odp(self, file_path: str) -> str:
        """Converte arquivo ODP para texto usando a extração de zip/xml."""
        try:
            return await self._extract_text_from_zip_xml(file_path, '', 'content.xml')
        except Exception as e:
            # Se a extração do zip falhar, tenta o método antigo como fallback
            try:
                return await self._convert_odp_fallback(file_path)
            except Exception as fallback_e:
                raise Exception(f"Falha na conversão de ODP com zip ({e}) e fallback ({fallback_e})")

    async def _convert_odp_fallback(self, file_path: str) -> str:
        """Fallback para conversão de ODP usando odfpy."""
        if load is None or P is None or extractText is None:
            raise ImportError("odfpy não está instalado para o fallback de ODP.")
        text_content = []
        doc = load(file_path)
        for p in doc.getElementsByType(P):
            text_content.append(extractText(p))
        return '\n'.join(text_content)
    
    async def _convert_ods(self, file_path: str) -> str:
        """Converte arquivo ODS para texto"""
        if pd is None:
            raise ImportError("pandas não está instalado")
        
        # Lê todas as planilhas do arquivo ODS
        sheets = pd.read_excel(file_path, sheet_name=None, engine='odf')
        text_content = []
        
        for sheet_name, df in sheets.items():
            text_content.append(f"=== Planilha: {sheet_name} ===")
            
            # Converte DataFrame para string
            csv_string = df.to_csv(index=False, sep='\t')
            text_content.append(csv_string)
        
        return '\n'.join(text_content)
    
    async def _convert_json(self, file_path: str) -> str:
        """Converte arquivo JSON para texto"""
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        return json.dumps(data, indent=2, ensure_ascii=False)
    
    def _check_libreoffice_availability(self) -> bool:
        """Verifica se o LibreOffice está disponível no sistema."""
        try:
            result = subprocess.run(['libreoffice', '--version'], 
                                  capture_output=True, text=True, timeout=5)
            return result.returncode == 0
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return False

    def _check_antiword_availability(self) -> bool:
        """Verifica se o antiword está disponível no sistema."""
        try:
            result = subprocess.run(['antiword', '-v'], 
                                  capture_output=True, text=True, timeout=5)
            return result.returncode == 0
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return False

    def _check_catdoc_availability(self) -> bool:
        """Verifica se o catdoc está disponível no sistema."""
        try:
            result = subprocess.run(['catdoc', '-V'], 
                                  capture_output=True, text=True, timeout=5)
            return result.returncode == 0
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return False

    def _convert_doc_with_antiword(self, file_path: str) -> str:
        """Converte arquivo DOC usando antiword."""
        try:
            result = subprocess.run(['antiword', '-t', file_path], 
                                  capture_output=True, text=True, timeout=30)
            if result.returncode == 0:
                return result.stdout
            else:
                raise Exception(f"antiword falhou com código {result.returncode}: {result.stderr}")
        except subprocess.TimeoutExpired:
            raise Exception("Timeout ao executar antiword")
        except Exception as e:
            raise Exception(f"Erro ao executar antiword: {e}")

    def _convert_doc_with_catdoc(self, file_path: str) -> str:
        """Converte arquivo DOC usando catdoc."""
        try:
            result = subprocess.run(['catdoc', '-a', file_path], 
                                  capture_output=True, text=True, timeout=30)
            if result.returncode == 0:
                return result.stdout
            else:
                raise Exception(f"catdoc falhou com código {result.returncode}: {result.stderr}")
        except subprocess.TimeoutExpired:
            raise Exception("Timeout ao executar catdoc")
        except Exception as e:
            raise Exception(f"Erro ao executar catdoc: {e}")

    async def _convert_doc(self, file_path: str) -> str:
        """Converte arquivo DOC para texto usando antiword ou catdoc como fallback."""
        errors = []
        
        # Tenta antiword primeiro
        if self._antiword_available:
            try:
                text = self._convert_doc_with_antiword(file_path)
                # antiword já produz texto limpo, apenas normalizar espaços
                normalized_text = re.sub(r'\s+', ' ', text.strip())
                return self.clean_text(normalized_text)
            except Exception as e:
                errors.append(f"antiword: {e}")
        
        # Tenta catdoc como fallback
        if self._check_catdoc_availability():
            try:
                text = self._convert_doc_with_catdoc(file_path)
                normalized_text = re.sub(r'\s+', ' ', text.strip())
                return self.clean_text(normalized_text)
            except Exception as e:
                errors.append(f"catdoc: {e}")
        
        # Se nenhuma ferramenta funcionou
        if not self._antiword_available and not self._check_catdoc_availability():
            raise Exception("Nenhuma ferramenta de conversão .doc está disponível (antiword ou catdoc)")
        
        # Se as ferramentas estão disponíveis mas falharam
        error_msg = "Falha ao converter arquivo .doc. Erros: " + "; ".join(errors)
        raise Exception(error_msg)
    
    async def _convert_xls(self, file_path: str) -> str:
        """Converte arquivo XLS para texto usando pandas."""
        if pd is None:
            raise ImportError("pandas não está instalado. Não é possível converter arquivos .xls")

        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
            text_content = []
            
            for sheet_name, df in sheets.items():
                text_content.append(f"=== Planilha: {sheet_name} ===")
                csv_string = df.to_csv(index=False, sep='\t')
                text_content.append(csv_string)
            
            return '\n'.join(text_content)
        except Exception as e:
            raise Exception(f"Falha ao converter .xls com pandas: {e}")